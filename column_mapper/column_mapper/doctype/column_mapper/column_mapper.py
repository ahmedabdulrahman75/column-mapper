# Copyright (c) 2025, Ahmed AbdulRahman and contributors
# For license information, please see license.txt

import frappe
import secrets
from typing import TypedDict
from typing_extensions import NotRequired
from io import BytesIO
from PIL import Image as Img
from frappe.model.document import Document
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.drawing.image import Image


class FormattedMapper(TypedDict):
	target_field: str
	field_type: str
	idx: NotRequired[int]


class ColumnMapper(Document):
	def __init__(self, *args, **kwargs):
		super().__init__(*args, **kwargs)
		self._file_doc: Document | None = None

	@property
	def file_doc(self) -> Document:
		if self._file_doc:
			return self._file_doc

		self._file_doc = frappe.get_doc("File", {"file_url": self.source_file})
		return self._file_doc

	def save_data(self):
		intersected_fields = self._get_intersected_fields_with_file()
		if not intersected_fields:
			return None

		ws = self._read_file()
		rows = ws.iter_rows(values_only=True)
		next(rows)
		for row_idx, row in enumerate(rows, start=1):
			record = self._create_record(intersected_fields, row, row_idx)
			if self.document_type == "Table":
				self.insert_into_table(record)
			else:
				self.insert_into_child_table(record)

	def upload_file(self, content, img_format) -> str:
		file_doc = frappe.new_doc("File")
		file_doc.file_name = f"new_file{secrets.token_urlsafe(8)}.{img_format}"
		file_doc.file_type = img_format
		file_doc.content = content
		file_doc.is_private = False
		file_doc.save()
		return file_doc.file_url

	def insert_into_table(self, record):
		doc = frappe.new_doc(self.ref_doctype)
		doc.update(record)
		doc.save()

	def insert_into_child_table(self, record):
		doc = frappe.get_cached_doc(self.ref_doctype, self.ref_doc)
		doc.append(self.doc_field, record)
		doc.save()

	def _get_intersected_fields_with_file(self) -> list[FormattedMapper]:
		file_columns = self._get_file_columns()
		mapper = self._get_formatted_field_mapper()
		mapped_columns: list = []
		for idx, col in enumerate(file_columns):
			if col in mapper:
				mapped_columns.append(
					{
						"idx": idx,
						"field_type": mapper[col]["field_type"],
						"target_field": mapper[col]["target_field"],
					}
				)

		return mapped_columns

	def _create_record(
		self,
		intersected_fields: list[FormattedMapper],
		row: list[str],
		row_idx: int,
	) -> dict:
		record = {}

		for field_data in intersected_fields:
			data = None
			if field_data["field_type"] != "Image":
				data = row[field_data["idx"]]
			else:
				img = self._get_image_content(row_idx, field_data["idx"])
				if img:
					data = self.upload_file(img[0], img[1])

			record[field_data["target_field"]] = data

		return record

	def _get_file_columns(self) -> list:
		ws = self._read_file()
		return [col.value for col in ws[1]]

	def _get_formatted_field_mapper(self) -> dict[str, FormattedMapper]:
		mapper: dict[str, FormattedMapper] = {}
		for field in self.field_mapper:
			mapper[field.file_field] = {
				"target_field": field.target_field,
				"field_type": field.field_type,
			}
		return mapper

	def _get_image_content(self, row, col) -> tuple[bytes, str] | None:
		img = self._get_image(row, col)
		if not img:
			return None

		buffer: BytesIO = img.ref
		buffer.seek(0)
		pil_img = Img.open(buffer)
		return img.ref.getvalue(), pil_img.format

	def _get_image(self, row, col) -> Image | None:
		ws = self._read_file()
		for img in ws._images:
			print("image data = ", img.anchor._from)
			if img.anchor._from.row == row and img.anchor._from.col == col:
				return img
		return None

	def _read_file(self) -> Worksheet:
		wb = load_workbook(self.file_doc.get_full_path(), data_only=True)
		ws = wb.active

		return ws

	@frappe.whitelist()
	def import_data(self):
		self.save_data()
